import pandas as pd
from fpdf import FPDF
import glob
import openpyxl
from pathlib import Path

#Will return a list object with all .xlsx files in the folder.
filepaths = glob.glob("invoices/*.xlsx")


for path in filepaths:
    data = pd.read_excel(path, sheet_name="Sheet 1")
    name, date = Path(path).stem.split("-")
    header = data.columns

    # Don't quite understand this syntax still...
    header = [item.replace("_", " ").title() for item in header]

    pdf = FPDF(orientation="portrait", unit="mm", format="letter")
    pdf.add_page()

    pdf.set_font(family="Arial", size=20, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice #{name}", ln=1)

    pdf.set_font(family="Arial", size=12, style="i")
    pdf.cell(w=50, h=8, txt=f"Date: {date}", ln=1)

    pdf.ln(25)

    pdf.set_font(family="Arial", size=10, style="B")

    pdf.cell(w=25, h=8, txt=header[0], border=1)
    pdf.cell(w=70, h=8, txt=header[1], border=1)
    pdf.cell(w=40, h=8, txt=header[2], border=1)
    pdf.cell(w=30, h=8, txt=header[3], border=1)
    pdf.cell(w=30, h=8, txt=header[4], border=1, ln=1)

    for index, row in data.iterrows():
        pdf.set_font(family="Arial", size=8)

        pdf.cell(w=25, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=40, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    total = data["total_price"].sum()
    pdf.cell(w=25, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=40, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(total), border=1, ln=1)

    pdf.ln(15)

    pdf.set_font(family="arial", size=12, style="i")
    pdf.cell(w=60, h=8, txt=f"The amount owed is {total}.")
    pdf.image("pythonhow.png", w=10)


    pdf.output(f"PDFs/{name}.pdf")